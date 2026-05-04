"""
Excel export funksiyasini oddiy test qilish (configga bog'liqmas)
"""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from aiogram.types import BufferedInputFile


# Mock data
class MockShipment:
    def __init__(self, cargo_id, full_name, phone, description, weight, cargo_weight, price, currency, status, notes, created_at):
        self.client = MockClient(cargo_id, full_name, phone)
        self.description = description
        self.weight_kg = weight
        self.cargo_weight_kg = cargo_weight
        self.price = price
        self.currency = currency
        self.status = status
        self.notes = notes
        self.photo_file_id = None
        self.created_at = created_at


class MockClient:
    def __init__(self, cargo_id, full_name, phone):
        self.cargo_id = cargo_id
        self.full_name = full_name
        self.phone_number = phone


def generate_excel_file_test(shipments):
    """
    Excel fayl generatsiya qiladi (test versiyasi)
    """
    wb = Workbook()
    ws = wb.active

    # Headers
    headers = [
        "📦", "Cargo ID", "👤", "Mijoz ismi", "📞", "Telefon",
        "📋", "Yuk nomi", "⚖️", "Kilo", "📦", "Cargo kilosi", "💰", "Narxi", "Valyuta", "📌", "Holat", "📅", "Yaratilgan"
    ]

    header_fill = PatternFill(
        start_color="00AEEFFF",
        end_color="00AEEFFF",
        fgColor="FF000000"
    )

    header_font = Font(bold=True)

    # Header qo'shish
    for col_num, header_text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header_text
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = header_font

    # Column widths
    column_widths = {
        "A": 8, "B": 15, "C": 8, "D": 20, "E": 8, "F": 15,
        "G": 8, "H": 30, "I": 8, "J": 10, "K": 8, "L": 12,
        "M": 8, "N": 10, "O": 8, "P": 8, "Q": 15, "R": 8, "S": 20
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Data qo'shish
    row_num = 2
    for shipment in shipments:
        status_emoji_map = {
            "pending": "🕐",
            "in_transit": "🚚",
            "arrived": "📦",
            "ready": "✅",
            "delivered": "🎉",
        }
        status_emoji = status_emoji_map.get(shipment.status, "")

        ws.cell(row=row_num, column=1, value=row_num - 1)
        ws.cell(row=row_num, column=2, value=shipment.client.cargo_id)
        ws.cell(row=row_num, column=3, value=shipment.client.full_name or "")
        ws.cell(row=row_num, column=4, value=shipment.client.phone_number)
        ws.cell(row=row_num, column=5, value=shipment.description or "-")
        ws.cell(row=row_num, column=6, value=shipment.weight_kg or "")
        ws.cell(row=row_num, column=7, value=shipment.cargo_weight_kg or "")
        ws.cell(row=row_num, column=8, value=f"{shipment.price}" if shipment.price else "")
        ws.cell(row=row_num, column=9, value=shipment.currency or "")
        ws.cell(row=row_num, column=10, value=status_emoji)
        ws.cell(row=row_num, column=11, value=shipment.photo_file_id or "")
        ws.cell(row=row_num, column=12, value=shipment.created_at.strftime('%d.%m.%Y %H:%M') if shipment.created_at else "")
        ws.cell(row=row_num, column=13, value=shipment.notes or "")

        row_num += 1

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"cargo_export_{datetime.now().strftime('%Y-%m-%d')}.xlsx"

    return BufferedInputFile(
        file=buffer.read(),
        filename=filename
    )


def test_excel():
    """Excel export testi"""
    print("[TEST] Excel export testi boshlandi...")

    # Mock shipments
    shipments = [
        MockShipment(
            cargo_id="12345",
            full_name="Alijon Valiyev",
            phone="+998901234567",
            description="Kiyim-kechak",
            weight=25.5,
            cargo_weight=30.0,
            price=450.00,
            currency="USD",
            status="in_transit",
            notes="Ehtiyotkorlik bilan",
            created_at=datetime.now()
        ),
        MockShipment(
            cargo_id="67890",
            full_name="Jamshid Karimov",
            phone="+998909876543",
            description="Elektronika",
            weight=10.0,
            cargo_weight=15.0,
            price=200.00,
            currency="USD",
            status="pending",
            notes=None,
            created_at=datetime.now()
        ),
    ]
    print(f"[TEST] {len(shipments)} ta mock shipment yaratildi")

    try:
        result = generate_excel_file_test(shipments)

        print(f"[OK] Excel fayl muvaffaqiyatli generatsiya qilindi!")
        print(f"[FILE] Fayl nomi: {result.filename}")

        # Faylni diskka saqlash
        filename = f"test_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        with open(filename, "wb") as f:
            f.write(result.read())

        print(f"[SAVE] Fayl diskka saqlandi: {filename}")
        print("\n[SUCCESS] Barcha testlar muvaffaqiyatli o'tdi!")

    except Exception as e:
        print(f"[ERROR] Xatolik yuz berdi: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    test_excel()
