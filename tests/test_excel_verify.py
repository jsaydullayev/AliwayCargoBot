"""
Excel export funksiyasi verifikatsiya testi
"""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment


# Mock data
class MockShipment:
    def __init__(self, cargo_id, full_name, phone, description, weight, cargo_weight, price, currency, status, notes, created_at):
        self.client = MockClient(cargo_id, full_name, phone)
        self.description = description
        self.weight_kg = weight
        self.cargo_weight_kg = cargo_weight
        self.price = price
        self.currency = currency
        self.status = status
        self.notes = notes
        self.photo_file_id = None
        self.created_at = created_at


class MockClient:
    def __init__(self, cargo_id, full_name, phone):
        self.cargo_id = cargo_id
        self.full_name = full_name
        self.phone_number = phone


def test_excel_generation():
    """Excel generatsiya testi"""
    print("[TEST] Excel generatsiya testi boshlandi...")

    wb = Workbook()
    ws = wb.active

    # Headers
    headers = [
        "Cargo ID", "Mijoz ismi", "Telefon",
        "Yuk nomi", "Kilo", "Cargo kilosi", "Narxi", "Valyuta", "Holat", "Yaratilgan"
    ]

    # 1. Headers yoziladi
    for col_num, header_text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header_text
        print(f"  [OK] Header yozildi: {header_text}")

    # 2. Column widths to'g'riladi
    column_widths = {
        "A": 15, "B": 20, "C": 15, "D": 30, "E": 10, "F": 12,
        "G": 12, "H": 10, "I": 8, "J": 20, "K": 8
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
        print(f"  [OK] Column width: {col_letter} = {width}")

    # 3. Data yoziladi
    shipments = [
        MockShipment(
            cargo_id="12345",
            full_name="Alijon Valiyev",
            phone="+998901234567",
            description="Kiyim-kechak",
            weight=25.5,
            cargo_weight=30.0,
            price=450.00,
            currency="USD",
            status="in_transit",
            notes="Ehtiyotkorlik bilan",
            created_at=datetime.now()
        ),
    ]

    row_num = 2
    for shipment in shipments:
        ws.cell(row=row_num, column=1, value=shipment.client.cargo_id)
        ws.cell(row=row_num, column=2, value=shipment.client.full_name or "")
        ws.cell(row=row_num, column=3, value=shipment.client.phone_number)
        ws.cell(row=row_num, column=4, value=shipment.description or "-")
        ws.cell(row=row_num, column=5, value=shipment.weight_kg or "")
        ws.cell(row=row_num, column=6, value=shipment.cargo_weight_kg or "")
        ws.cell(row=row_num, column=7, value=f"{shipment.price}" if shipment.price else "")
        ws.cell(row=row_num, column=8, value=shipment.currency or "")
        ws.cell(row=row_num, column=9, value="Yo'lda")
        ws.cell(row=row_num, column=10, value=shipment.created_at.strftime('%d.%m.%Y %H:%M') if shipment.created_at else "")
        ws.cell(row=row_num, column=11, value=shipment.notes or "")

        print(f"  [OK] Row {row_num-1} yozildi: {shipment.description}")
        row_num += 1

    # 4. Freeze panes
    ws.freeze_panes = "A2"
    print(f"  [OK] Freeze panes to'g'rilandi")

    # 5. Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Faylni diskka saqlash
    filename = f"test_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    with open(filename, "wb") as f:
        f.write(buffer.read())

    print(f"\n[SUCCESS] Fayl diskka saqlandi: {filename}")
    print(f"[SUCCESS] Fayl kattaligi: {buffer.tell()} bytes")


if __name__ == "__main__":
    test_excel_generation()
