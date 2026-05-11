"""
Excel eksport (TZ §5.6)

Ustun tartibi: Cargo ID | Mijoz ismi | Telefon | Yuk nomi | Kilo |
              Cargo kilo | Narxi | Valyuta | Status | Yaratilgan
"""
import io
from datetime import datetime
from typing import List, Any

from aiogram.types import BufferedInputFile

from database.models import CargoStatus


STATUS_LABELS = {
    CargoStatus.PENDING: "🕐 Pending",
    CargoStatus.IN_TRANSIT: "🚚 In transit",
    CargoStatus.ARRIVED: "📦 Arrived",
    CargoStatus.READY: "✅ Ready",
    CargoStatus.DELIVERED: "🎉 Delivered",
}


def generate_excel_file(shipments: List[Any]) -> BufferedInputFile:
    """
    Yuk yozuvlaridan .xlsx fayl generatsiya qiladi.

    Args:
        shipments: ORM Shipment obyektlari yoki tuple ro'yxati

    Returns:
        BufferedInputFile — aiogram'da `answer_document` ga uzatish uchun tayyor
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Cargo Export"

    headers = [
        "Cargo ID",
        "Mijoz ismi",
        "Telefon",
        "Yuk nomi",
        "Kilo (kg)",
        "Cargo kilo (kg)",
        "Narxi",
        "Valyuta",
        "Status",
        "Yaratilgan",
    ]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    column_widths = [12, 22, 18, 30, 12, 16, 12, 10, 18, 20]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    for row_idx, shipment in enumerate(shipments, start=2):
        # ORM Shipment'da `.client` bor, SQLAlchemy Row'da yo'q
        if hasattr(shipment, "client"):
            cargo_id = shipment.client.cargo_id if shipment.client else "—"
            full_name = (shipment.client.full_name if shipment.client else "") or ""
            phone_number = shipment.client.phone_number if shipment.client else ""
            description = shipment.description or ""
            weight_kg = shipment.weight_kg
            cargo_weight_kg = shipment.cargo_weight_kg
            price = shipment.price
            currency = shipment.currency or ""
            status = shipment.status
            created_at = shipment.created_at
        else:
            cargo_id, full_name, phone_number, description, weight_kg, \
                cargo_weight_kg, price, currency, status, created_at = tuple(shipment)

        status_label = STATUS_LABELS.get(status, str(status.value) if status else "")
        created_str = created_at.strftime("%d.%m.%Y %H:%M") if created_at else ""

        row_values = [
            cargo_id or "—",
            full_name,
            phone_number,
            description,
            float(weight_kg) if weight_kg else None,
            float(cargo_weight_kg) if cargo_weight_kg else None,
            float(price) if price else None,
            currency,
            status_label,
            created_str,
        ]

        for col_num, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_num, value=value)
            cell.border = thin_border
            if col_num in (5, 6, 7):  # numeric columns
                cell.number_format = "#,##0.00"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"cargo_export_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return BufferedInputFile(file=buffer.read(), filename=filename)


LANGUAGE_LABELS = {
    "uz": "🇺🇿 O'zbek",
    "ru": "🇷🇺 Русский",
    "tr": "🇹🇷 Türkçe",
}


def generate_clients_excel_file(clients: List[Any]) -> BufferedInputFile:
    """
    Mijozlar ro'yxatidan .xlsx fayl generatsiya qiladi.

    Args:
        clients: tuple ro'yxati (cargo_id, full_name, phone, telegram_id, language, created_at, shipments_count)

    Returns:
        BufferedInputFile — aiogram'da `answer_document` ga uzatish uchun tayyor
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Clients"

    headers = [
        "#",
        "Cargo ID",
        "Mijoz ismi",
        "Telefon",
        "Telegram",
        "Til",
        "Yuklar soni",
        "Ro'yxatdan o'tgan",
    ]

    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    column_widths = [5, 12, 25, 18, 16, 14, 12, 20]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    for row_idx, client in enumerate(clients, start=2):
        cargo_id, full_name, phone, telegram_id, language, created_at, ship_count = tuple(client)

        lang_label = LANGUAGE_LABELS.get(language, language or "—")
        telegram_str = str(telegram_id) if telegram_id else "—"
        created_str = created_at.strftime("%d.%m.%Y %H:%M") if created_at else ""

        row_values = [
            row_idx - 1,
            cargo_id or "—",
            full_name or "",
            phone or "",
            telegram_str,
            lang_label,
            int(ship_count or 0),
            created_str,
        ]

        for col_num, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_num, value=value)
            cell.border = thin_border
            if col_num in (1, 7):  # # va yuklar soni — markazga
                cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"clients_export_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return BufferedInputFile(file=buffer.read(), filename=filename)
